import os
import openpyxl
from openpyxl.styles import PatternFill
import colour

# Get the file name of the only available file in the input directory
input_dir = os.path.join(os.path.abspath('.'), 'input')
file_name = [f for f in os.listdir(input_dir) if os.path.isfile(os.path.join(input_dir, f))][0]

# Open the Excel file
wb = openpyxl.load_workbook(os.path.join(input_dir, file_name))

# Read the settings from the settings.txt file
with open('settings.txt', 'r') as f:
    settings = {line.strip().split('=')[0]: line.strip().split('=')[1] for line in f.readlines()}

# Convert color names to ARGB hex values
settings = {k: colour.Color(v).get_hex_l()[1:] for k, v in settings.items()}

# Print all available worksheets
print("Dostupne radne knjige:")
for i, sheet in enumerate(wb.worksheets):
    print(f' \033[92m {i + 1}\033[0m. {sheet.title}')


# Prompt the user to select a worksheet
sheet_index = int(input('Molimo upišite broj radne knjige koju želite formatirati: ')) - 1
ws = wb.worksheets[sheet_index]
print('Odabrana radna knjiga za formatiranje je: \033[92m"' + ws.title + '"\033[0m')

# Determine the range of cells
min_row = '1'
min_col = '1'
max_row = ws.max_row
max_col = ws.max_column

start_cell = f'{chr(ord(min_col) + int(min_col))}{min_row}'
end_cell = f'{chr(ord(min_col) + max_col)}{max_row}'

cell_range = f'{start_cell}:{end_cell}'

# Apply the conditional formatting to the data range
for row in range(int(min_row), max_row + 1):
    for col in range(int(min_col), max_col + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value in settings:
            # print(f'Cell {cell.coordinate} has value {cell.value} and will be formatted with color {settings[cell.value]}')
            cell.fill = PatternFill(start_color=settings[cell.value], fill_type='solid')

# Save the workbook to the output directory
output_dir = os.path.join(os.path.abspath('.'), 'output')
os.makedirs(output_dir, exist_ok=True)
output_file_name = os.path.join(output_dir, file_name)
wb.save(output_file_name)

print(f'Radna knjiga je uspješno formatirana i spremljena u: {output_file_name}')